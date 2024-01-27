from openpyxl import load_workbook

# Updated list of fuels
fuel_list = [
    "EULLND", "EULMAI", "EULWHE", "EULBAR", "EULSUN", "EULRAP", "EULSOY", "EULOAT",
    "EULOLI", "EULRHY", "EULPOT", "EULOTH", "EXLMAI", "EXLWHE", "EXLBAR", "EXLSUN",
    "EXLRAP", "EXLSOY", "EXLOAT", "EXLOLI", "EXLRHY", "EXLOTP", "EXLOTH", "EULFOR",
    "EULBLT", "EULWAT", "EULBAR", "EULSNO", "EULMMW", "EULGRS", "EULPAS",
    "EULLIV", "EULMEA", "EULADSL", "EULTDSL", "EULTETH", "EUWPRC", "EUWAGR", "EUWPUB",
    "EUWOTH", "EUWPWR", "EUWDAG", "EUWDPU", "EUWDPW", "EUWDOT", "EUWGWT", "EUWSUR",
    "EUWSEA", "EUWEVT", "EUWCLE"
]

# Define the path to your Excel file
file_path = r"C:\Users\leigh\Documents\Otool CLEWsEU\template.xlsx"

# Load the Excel file using openpyxl
workbook = load_workbook(filename=file_path)

# Sheets to exclude from updating
exclude_sheets = ['OutputActivityRatio', 'InputActivityRatio']

# Iterate over all sheets in the workbook, excluding specified ones
for sheet_name in workbook.sheetnames:
    if sheet_name not in exclude_sheets:
        sheet = workbook[sheet_name]

        # Find the 'FUEL' column index, assuming it's in the first row
        fuel_column_index = None
        for i, cell in enumerate(sheet[1]):
            if cell.value == 'FUEL':
                fuel_column_index = i + 1
                break

        # If 'FUEL' column is present, delete its contents (except for the first row)
        if fuel_column_index:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=fuel_column_index).value = None

            # Populate 'FUEL' column with the updated list
            for row_index, fuel in enumerate(fuel_list, start=2):
                sheet.cell(row=row_index, column=fuel_column_index, value=fuel)

# Save the workbook
workbook.save(filename=file_path)

# Output success message
print(f"Fuels have been updated in the Excel file: {file_path}")