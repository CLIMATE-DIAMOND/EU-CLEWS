from openpyxl import load_workbook

# List of technologies
technology_list = [
    "EUL000MIN000", "EULIMPMAI000", "EULIMPWHE000", "EULIMPBAR000",
    "EULIMPSUN000", "EULIMPRAP000", "EULIMPSOY000", "EULIMPOAT000",
    "EULIMPOLI000", "EULIMPRHY000", "EULIMPPOT000", "EULIMPOTH000",
    "EULEXPMAI000", "EULEXPWHE000", "EULEXPBAR000", "EULEXPSUN000",
    "EULEXPRAP000", "EULEXPSOY000", "EULEXPOAT000", "EULEXPOLI000",
    "EULEXPRHY000", "EULEXPPOT000", "EULEXPOTH000", "EUL000MAIHR0",
    "EUL000WHEHR0", "EUL000BARHR0", "EUL000SUNHR0", "EUL000RAPHR0",
    "EUL000SOYHR0", "EUL000OATHR0", "EUL000OLIHR0", "EUL000RHYHR0",
    "EUL000POTHR0", "EUL000OTHHR0", "EUL000MAIHI0", "EUL000WHEHI0",
    "EUL000BARHI0", "EUL000SUNHI0", "EUL000RAPHI0", "EUL000SOYHI0",
    "EUL000OATHI0", "EUL000OLIHI0", "EUL000RHYHI0", "EUL000POTHI0",
    "EUL000OTHHI0", "EUL000FOR000", "EUL000BLT000", "EUL000WAT000",
    "EUL000BRN000", "EUL000SNO000", "EUL000MMW000", "EUL000GRS000",
    "EUL000CRP000", "EUL000PAS000", "EUL000LIV000", "EUL000MEA000",
    "EULADSL000", "EULTDSL000", "EULTETH000", "EUWMIN000PRC",
    "EUWDEMAGRSUR", "EUWDEMAGRGWT", "EUWDEMPWRSUR", "EUWDEMPWRGWT",
    "EUWDEMPUBSUR", "EUWDEMPUBGWT", "EUWDEMOTHSUR", "EUWDEMOTHGWT",
    "EUWTRNPUB000", "EUWTRNAGR000", "EUWTRNPWR000", "EUWTRNOTH000",
    "EUWDSA000000", "EUWMIN000SEA", "EUWTRN000TRE"
]

# Define the path to your Excel file (use a raw string for file paths)
file_path = r"C:\Users\leigh\Documents\Otool CLEWsEU\template.xlsx"

# Load the Excel file using openpyxl for more flexible editing
workbook = load_workbook(filename=file_path)

# Sheets to exclude
exclude_sheets = ['InputActivityRatio', 'OutputActivityRatio']

# Iterate over all sheets in the workbook
for sheet_name in workbook.sheetnames:
    if sheet_name not in exclude_sheets:
        sheet = workbook[sheet_name]
    
        # Assuming 'Technology' is in the first row, find the 'Technology' column index
        technology_column_index = None
        for i, cell in enumerate(sheet[1]):
            if cell.value == 'TECHNOLOGY':
                technology_column_index = i + 1
                break
    
        # If 'Technology' column is present in the sheet, populate it
        if technology_column_index:
            for row_index, technology in enumerate(technology_list, start=2):
                sheet.cell(row=row_index, column=technology_column_index, value=technology)

# Save the workbook
workbook.save(filename=file_path)

# Output a success message
print(f"Technologies have been populated in the Excel file: {file_path}")
