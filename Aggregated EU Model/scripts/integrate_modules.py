# Code to merge the Energy module with the Land and water Module to create the integrated_CLEWs module

import openpyxl

# Load the first workbook corresponding to the land and water module
wb1 = openpyxl.load_workbook(filename='land_water_module.xlsx')

# Load the second workbook corresponding to the energy module
wb2 = openpyxl.load_workbook(filename='energy_module.xlsx')

# List of sheet names to exclude (to avoid duplication)
exclude_sheets = ['YEAR', 'YearSplit','MODE_OF_OPERATION','TIMESLICE','REGION']  

# Loop over the sheets in the first workbook
for sheet_name in wb1.sheetnames:
    # Skip the sheets in the exclusion list
    if sheet_name in exclude_sheets:
        continue

    # Get the sheet from the first workbook
    sheet1 = wb1[sheet_name]

    # Check if the sheet exists in the second workbook
    if sheet_name in wb2.sheetnames:
        # Get the sheet from the second workbook
        sheet2 = wb2[sheet_name]

        # Get the maximum row in the sheet from the first workbook
        max_row = sheet1.max_row

        # Loop over the rows in the sheet from the second workbook
        for row in sheet2.iter_rows(min_row=2, values_only=True):
            # Append the row to the sheet from the first workbook
            sheet1.append(row)

# Save the first workbook with a new name. Make sure to included the (.xlsx) extension 
wb1.save('integrated_clews_version.xlsx')
